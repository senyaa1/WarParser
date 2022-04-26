import requests
import json
from openpyxl import Workbook


def safe_add(ws, row, col, val):
    if val is None:
        val = "Нет данных"
    ws.cell(row, col, val)


url = "https://gwar.mil.ru/gt_data/?builder=Heroes"

size = input("Максимальное количество результатов: ")
table_path = input("Путь сохранения (table.xlsx): ")

first_name = input("Имя: ")
middle_name = input("Отчество: ")
last_name = input("Фамилия: ")
birth_date = input("Дата рождения (ДД.ММ.ГГГГ): ")
birth_place = input("Место рождения: ")
birth_place_gubernia = input("Место рождения (Губерния): ")
birth_place_uezd = input("Место рождения (Уезд): ")
birth_place_volost = input("Место рождения (Волость): ")


body = """{
    "indices": [
        "gwar"
    ],
    "entities": [
        "chelovek_donesenie",
        "chelovek_gospital",
        "chelovek_zahoronenie",
        "chelovek_plen",
        "chelovek_nagrazhdenie",
        "chelovek_predstavlenie",
        "chelovek_nagradnaya_kartochka",
        "commander",
        "person",
        "chelovek_posluzhnoi_spisok",
        "chelovek_uchetnaya_kartochka"
    ],
    "queryFields": {
        "ids": "",
        "last_name": "LASTNAME",
        "first_name": "FIRSTNAME",
        "middle_name": "MIDDLENAME",
        "birth_place": "BIRTHLOCATION",
        "birth_place_gubernia": "BIRTHPLACEGUBERNIA",
        "birth_place_uezd": "BIRTHPLACEUEZD",
        "birth_place_volost": "BIRTHPLACEVOLOST",
        "location": "",
        "birth_date": "",
        "rank": "",
        "data_vibitiya": "",
        "event_name": "",
        "event_id": "",
        "military_unit_name": "",
        "event_place": "",
        "lazaret_name": "",
        "camp_name": "",
        "date_death": "",
        "award_name": "",
        "nomer_dokumenta": "",
        "data_dokumenta": "",
        "data_i_mesto_priziva": "",
        "archive_short": "",
        "nomer_fonda": "",
        "nomer_opisi": "",
        "nomer_dela": "",
        "date_birth": "",
        "data_vibitiya_end": ""
    },
    "filterFields": {},
    "from": 0,
    "size": "SIZE",
    "builderType": "Heroes"
}"""

body = body.replace("SIZE", size).replace("LASTNAME", last_name).replace("MIDDLENAME", middle_name).replace("FIRSTNAME", first_name).replace("BIRTHLOCATION", birth_place).replace(
    "BIRTHPLACEGUBERNIA", birth_place_gubernia).replace("BIRTHPLACEUEZD", birth_place_uezd).replace("BIRTHPLACEVOLOST", birth_place_volost)

print(f"Запрос на {url}...")
req = requests.post(url, data=body.encode('utf-8'))
result = json.loads(req.content.decode("unicode-escape"))

wb = Workbook()
ws = wb.active

headers = ["Фамилия", "Имя", "Отчество",
           "Дата рождения", "Место рождения", "Место службы", "Причина выбытия"]

for i, header in enumerate(headers):
    ws.cell(1, i + 1, header)


for i, entry in enumerate(result["hits"]["hits"]):
    try:
        safe_add(ws, i + 2, 1, entry["_source"]["last_name"])
        safe_add(ws, i + 2, 2, entry["_source"]["first_name"])
        safe_add(ws, i + 2, 3, entry["_source"]["middle_name"])
        safe_add(ws, i + 2, 4, entry["_source"]["birth_date"])
        safe_add(ws, i + 2, 5, entry["_source"]["birth_place"])
        safe_add(ws, i + 2, 6, entry["_source"]["military_unit_name"])
        safe_add(ws, i + 2, 7, entry["_source"]["vibitie_prichina"])
    except:
        continue

if table_path == "":
    table_path = "table.xlsx"
    
wb.save(table_path)

print(
    f"Таблица {table_path} с {len(result['hits']['hits'])} строками сохранена!")
